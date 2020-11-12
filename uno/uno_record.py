import openpyxl
from uno import uno_func


EXCEL_PASS = 'uno_record.xlsx'


def calculate_point(card):
    pts = 0
    for i in card:
        id = uno_func.card_to_id(i)
        if id % 100 <= 9 and id < 500:  # 数字カードはその数字の点数
            pts -= id % 100
        elif id < 500:  # 記号カードは20点
            pts -= 20
        else:  # ワイルドカードは50点
            pts -= 50

    return pts


def add_penalty(player, card):
    book = openpyxl.load_workbook(EXCEL_PASS)
    sheet1 = book['Records']
    sheet2 = book['History']
    pts = calculate_point(card)

    for j in range(1, sheet1.max_row + 1):
        # 既存ユーザーのデータ上書き
        if str(player) == sheet1.cell(2, j).value:
            # ポイント書き込み
            for k in range(2, sheet2.max_column + 1):
                if sheet1.cell(k, j).value == "":
                    sheet2.cell(k, j, pts - 100)
                    break
            # Lose 書き込み
            sheet1.cell(6, j, sheet1.cell(6, j).value + 1)
            break
        # 新規ユーザーのデータ作成
        elif sheet1.cell(2, j).value == "":
            # ID書き込み
            sheet1.cell(2, j, str(player))
            sheet2.cell(1, j, str(player))
            # Lose 書き込み
            sheet1.cell(5, j, 0)
            sheet1.cell(6, j, 1)
            # ペナルティー書き込み
            sheet1.cell(9, j, -100)
            # ポイント書き込み
            for k in range(2, sheet2.max_column + 1):
                if sheet2.cell(k, j).value == "":
                    sheet2.cell(k, j, pts - 100)
                    break
    book.save(EXCEL_PASS)


def data_save(data):
    book = openpyxl.load_workbook(EXCEL_PASS)
    sheet1 = book['Records']
    sheet2 = book['History']

    for i in range(len(data)):
        for j in range(1, sheet1.max_row + 1):
            # 既存ユーザーのデータ上書き
            if str(data[i][0]) == sheet1.cell(2, j).value:
                # ポイント書き込み
                for k in range(2, sheet2.max_column + 1):
                    if sheet1.cell(k, j).value == "":
                        sheet2.cell(k, j, data[i][4])
                        break
                # Win or Lose 書き込み
                if data[i][4] > 0:
                    sheet1.cell(5, j, sheet1.cell(5, j).value + 1)
                else:
                    sheet1.cell(6, j, sheet1.cell(6, j).value + 1)
                break
            # 新規ユーザーのデータ作成
            elif sheet1.cell(2, j).value == "":
                # ID書き込み
                sheet1.cell(2, j, str(data[i][0]))
                sheet2.cell(1, j, str(data[i][0]))
                # Win or Lose 書き込み
                if data[i][4] > 0:
                    sheet1.cell(5, j, 1)
                    sheet1.cell(6, j, 0)
                else:
                    sheet1.cell(5, j, 0)
                    sheet1.cell(6, j, 1)
                # ペナルティー書き込み
                sheet1.cell(9, j, 0)
                # ポイント書き込み
                for k in range(2, sheet2.max_column + 1):
                    if sheet2.cell(k, j).value == "":
                        sheet2.cell(k, j, data[i][4])
                        break
    book.save(EXCEL_PASS)


def data_delete(id):
    book = openpyxl.load_workbook(EXCEL_PASS)
    sheet1 = book['Records']
    sheet2 = book['History']

    for row in sheet1.iter_rows(min_row=2):
        if str(id) == row[0].values:
            sheet1.delete_rows(row)
            sheet2.delete_rows(row)
    book.save(EXCEL_PASS)


def record_output(id):
    book = openpyxl.load_workbook(EXCEL_PASS)
    sheet = book['Records']
    data = []
    for row in sheet.iter_rows(min_row=2):
        if str(id) == row[0].values:
            for col in row:
                data.append(col.value)
            break

    if not data:
        return None

    if data[2] <= 0:
        url = 'https://i.imgur.com/adtGl7h.png'  # YOU LOSE
    else:
        url = 'https://i.imgur.com/1JXc9eD.png'  # YOU WIN
        data[2] = f"+{data[2]}"
    if data[7] == "":
        data[7] = "N/A"
    if data[8] == "":
        data[8] = "N/A"
    if data[10] >= 0:
        data[10] = f"+{data[10]}"

    return data, url
